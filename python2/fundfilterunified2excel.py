# -*- coding:utf-8 -*-
import pandas as pd
import math
from datetime import date
from openpyxl import load_workbook


def formatDate(dt):
    format = "%Y-%m-%d"
    return dt.strftime(format)


todayStr = formatDate(date.today())

input = "/root/funddata/funddata_"+todayStr+".csv"
output = "/root/funddata/fund_unified_data_"+todayStr+".xlsx"
blacklistfile = "/root/funddata/blacklist.txt"
blacklist = set(open(blacklistfile).read().split())


def zhangfu(num1, num2):
    return (num2-num1)/num1


if __name__ == "__main__":
    iter_csv = pd.read_csv(input, encoding='gbk',
                           iterator=True, chunksize=1000)

    df = pd.concat([chunk[(chunk[u'当前净值'] > chunk[u'一年前净值'])
                          & (chunk[u'当前净值'] > chunk[u'三年前净值'])
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > zhangfu(chunk[u'三年前净值'], chunk[u'当前净值']))
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > 0.2)] for chunk in iter_csv])

    boolean_series = ~df[u'代码'].isin(blacklist)
    filtered_df = df[boolean_series]
    with pd.ExcelWriter(output) as writer:
        filtered_df.to_excel(writer, "Sheet_growfastover20",
                             encoding='utf-8-sig', index=False)
# =======分割线====
    iter_csv = pd.read_csv(input, encoding='gbk',
                           iterator=True, chunksize=1000)
    df = pd.concat([chunk[(zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > 1)]
                    for chunk in iter_csv])
    boolean_series = ~df[u'代码'].isin(blacklist)
    filtered_df = df[boolean_series]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        writer.book = load_workbook(output)
        filtered_df.to_excel(writer, "Sheet_oneyearcrazy",
                             encoding='utf-8-sig', index=False)
# =======分割线====
    iter_csv = pd.read_csv(input, encoding='gbk',
                           iterator=True, chunksize=1000)
    df = pd.concat([chunk[(chunk[u'当前净值'] > chunk[u'一年前净值'])
                          & (chunk[u'当前净值'] > chunk[u'三年前净值'])
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > zhangfu(chunk[u'三年前净值'], chunk[u'当前净值']))
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > 0.1)
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) < 0.2)] for chunk in iter_csv])
    boolean_series = ~df[u'代码'].isin(blacklist)
    filtered_df = df[boolean_series]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        writer.book = load_workbook(output)
        filtered_df.to_excel(writer, "Sheet_10_20_grow",
                             encoding='utf-8-sig', index=False)

# =======分割线====
    iter_csv = pd.read_csv(input, encoding='gbk',
                           iterator=True, chunksize=1000)
    df = pd.concat([chunk[(chunk[u'当前净值'] > chunk[u'一年前净值'])
                          & (chunk[u'一年前净值'] > chunk[u'两年前净值'])
                          & (chunk[u'两年前净值'] > chunk[u'三年前净值'])
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > zhangfu(chunk[u'三年前净值'], chunk[u'一年前净值']))
                          & (zhangfu(chunk[u'一年前净值'], chunk[u'当前净值']) > 0.2)] for chunk in iter_csv])
    boolean_series = ~df[u'代码'].isin(blacklist)
    filtered_df = df[boolean_series]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        writer.book = load_workbook(output)
        filtered_df.to_excel(writer, "Sheet_consec_grow_20",
                             encoding='utf-8-sig', index=False)
